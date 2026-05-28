from collections.abc import Generator
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import create_engine, event, select
from sqlalchemy.orm import Session, sessionmaker

from app.core.config import settings
from app.db.base import Base
from app.db.models.file_record import FileRecordModel
from app.db.models.task_record import TaskRecordModel
from app.db.session import get_db
from app.main import app


def _override_db_session(database_url: str) -> tuple[sessionmaker[Session], object]:
    engine = create_engine(
        database_url,
        future=True,
        connect_args={"check_same_thread": False},
    )

    @event.listens_for(engine, "connect")
    def _set_sqlite_pragma(dbapi_connection, connection_record) -> None:
        cursor = dbapi_connection.cursor()
        cursor.execute("PRAGMA foreign_keys=ON")
        cursor.close()

    Base.metadata.create_all(bind=engine)
    testing_session = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)

    def override_get_db() -> Generator[Session, None, None]:
        db = testing_session()
        try:
            yield db
        finally:
            db.close()

    return testing_session, override_get_db


def _build_sample_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "缁忚惀鍒嗘瀽"
    sheet["A1"] = "Header"
    sheet.merge_cells("A1:B1")
    sheet["A2"] = 100
    sheet["B2"] = 200
    workbook.save(path)
    workbook.close()


def _build_dual_track_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "鍙岃建"
    sheet["A1"] = "鍖哄煙"
    sheet.merge_cells("A1:A2")
    sheet["B1"] = 500
    sheet.merge_cells("B1:C1")
    sheet["B2"] = 100
    sheet["C2"] = 200
    workbook.save(path)
    workbook.close()


def _build_row_dimension_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "row-dimension"
    sheet["A1"] = "Region"
    sheet["B1"] = "Revenue"
    sheet["A2"] = "East"
    sheet["B2"] = 120
    workbook.save(path)
    workbook.close()


def _create_uploaded_task(session_factory: sessionmaker[Session], workbook_path: Path) -> int:
    with session_factory() as session:
        file_record = FileRecordModel(
            file_name=workbook_path.name,
            file_path=str(workbook_path),
            file_type=".xlsx",
        )
        session.add(file_record)
        session.flush()
        task_record = TaskRecordModel(file_id=file_record.id, status="uploaded")
        session.add(task_record)
        session.commit()
        session.refresh(task_record)
        return task_record.id


def test_review_returns_task_level_sheet_snapshots(tmp_path) -> None:
    session_factory, override_get_db = _override_db_session(
        f"sqlite:///{(tmp_path / 'test.db').as_posix()}"
    )
    app.dependency_overrides[get_db] = override_get_db
    settings.upload_dir = str(tmp_path / "uploads")
    client = TestClient(app)

    workbook_path = tmp_path / "source.xlsx"
    _build_sample_workbook(workbook_path)
    task_id = _create_uploaded_task(session_factory, workbook_path)

    parse_response = client.post(f"/api/tasks/{task_id}/parse")
    assert parse_response.status_code == 200

    response = client.get(f"/api/tasks/{task_id}/review")
    app.dependency_overrides.clear()

    assert response.status_code == 200
    payload = response.json()
    assert payload["task_id"] == task_id
    assert payload["status"] == "waiting_confirm"
    assert payload["structure_version"] == 0
    assert payload["editable_structure_version"] == 0
    assert len(payload["sheets"]) == 1

    sheet = payload["sheets"][0]
    assert sheet["sheet_name"] == "缁忚惀鍒嗘瀽"
    assert sheet["row_count"] == 2
    assert sheet["col_count"] == 2
    assert sheet["merge_ranges"] == ["A1:B1"]
    assert sheet["grid_snapshot"] == [["Header", None], ["100", "200"]]
    assert sheet["address_map"] == [["A1", "B1"], ["A2", "B2"]]
    assert sheet["aligned_grid"] == [["Header", "Header"], ["100", "200"]]
    assert sheet["aligned_cell_roles"] == [["dimension", "dimension"], ["measure", "measure"]]
    assert sheet["aligned_source_map"] == [["A1", "A1"], ["A2", "B2"]]
    assert sheet["header_row_span"] == 1
    assert sheet["column_paths"] == [["Header"], ["Header"]]
    assert sheet["column_kinds"] == ["dimension", "dimension"]
    assert sheet["dimension_columns"] == [0, 1]
    assert sheet["measure_columns"] == []
    assert sheet["raw_cells"][0]["address"] == "A1"
    assert sheet["raw_cells"][0]["merge_range"] == "A1:B1"


def test_review_returns_404_for_missing_task(tmp_path) -> None:
    _, override_get_db = _override_db_session(f"sqlite:///{(tmp_path / 'test.db').as_posix()}")
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    response = client.get("/api/tasks/999/review")
    app.dependency_overrides.clear()

    assert response.status_code == 404
    assert response.json()["detail"] == "Task not found"


def test_review_builds_dual_track_alignment_for_dimension_and_measure_merges(tmp_path) -> None:
    session_factory, override_get_db = _override_db_session(
        f"sqlite:///{(tmp_path / 'test.db').as_posix()}"
    )
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    workbook_path = tmp_path / "dual-track.xlsx"
    _build_dual_track_workbook(workbook_path)
    task_id = _create_uploaded_task(session_factory, workbook_path)

    parse_response = client.post(f"/api/tasks/{task_id}/parse")
    assert parse_response.status_code == 200

    response = client.get(f"/api/tasks/{task_id}/review")
    app.dependency_overrides.clear()

    assert response.status_code == 200
    payload = response.json()
    sheet = payload["sheets"][0]

    assert sheet["grid_snapshot"] == [["鍖哄煙", "500", None], [None, "100", "200"]]
    assert sheet["aligned_grid"] == [["鍖哄煙", "500", None], ["鍖哄煙", "100", "200"]]
    assert sheet["aligned_cell_roles"] == [
        ["dimension", "measure", "measure"],
        ["dimension", "measure", "measure"],
    ]
    assert sheet["aligned_source_map"] == [["A1", "B1", "C1"], ["A1", "B2", "C2"]]
    assert sheet["header_row_span"] == 2
    assert sheet["column_paths"] == [["鍖哄煙"], ["500", "100"], ["500", "200"]]
    assert sheet["column_kinds"] == ["dimension", "measure", "measure"]
    assert sheet["dimension_columns"] == [0]
    assert sheet["measure_columns"] == [1, 2]


def test_save_structure_version_persists_snapshot_and_review_prefers_latest_version(tmp_path) -> None:
    session_factory, override_get_db = _override_db_session(
        f"sqlite:///{(tmp_path / 'test.db').as_posix()}"
    )
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    workbook_path = tmp_path / "source.xlsx"
    _build_sample_workbook(workbook_path)
    task_id = _create_uploaded_task(session_factory, workbook_path)

    parse_response = client.post(f"/api/tasks/{task_id}/parse")
    assert parse_response.status_code == 200

    review_response = client.get(f"/api/tasks/{task_id}/review")
    assert review_response.status_code == 200
    review_payload = review_response.json()
    sheet = review_payload["sheets"][0]

    save_response = client.post(
        f"/api/tasks/{task_id}/structure-versions",
        json={
            "base_structure_version": 0,
            "sheets": [
                {
                    "sheet_id": sheet["sheet_id"],
                    "sheet_name": sheet["sheet_name"],
                    "sheet_index": sheet["sheet_index"],
                    "row_count": sheet["row_count"],
                    "col_count": sheet["col_count"],
                    "is_hidden": sheet["is_hidden"],
                    "merge_ranges": ["A1:B1", "A2:B2"],
                    "aligned_grid": [["Header", "Header"], ["North", "North"]],
                    "aligned_cell_roles": [["dimension", "dimension"], ["dimension", "dimension"]],
                    "aligned_source_map": [["A1", "A1"], ["A2", "A2"]],
                    "cell_tags": [["header", "header"], ["data", "data"]],
                }
            ],
        },
    )
    assert save_response.status_code == 201
    saved_payload = save_response.json()

    assert saved_payload["task_id"] == task_id
    assert saved_payload["structure_version"] == 1
    assert saved_payload["status"] == "waiting_confirm"
    assert saved_payload["patch_summary"]["sheet_count"] == 1
    assert saved_payload["patch_summary"]["changed_cell_count"] == 4

    latest_review_response = client.get(f"/api/tasks/{task_id}/review")
    app.dependency_overrides.clear()

    assert latest_review_response.status_code == 200
    latest_review_payload = latest_review_response.json()
    latest_sheet = latest_review_payload["sheets"][0]

    assert latest_review_payload["structure_version"] == 1
    assert latest_review_payload["editable_structure_version"] == 1
    assert latest_sheet["merge_ranges"] == ["A1:B1", "A2:B2"]
    assert latest_sheet["aligned_grid"] == [["Header", "Header"], ["North", "North"]]
    assert latest_sheet["aligned_cell_roles"] == [
        ["dimension", "dimension"],
        ["dimension", "dimension"],
    ]
    assert latest_sheet["aligned_source_map"] == [["A1", "A1"], ["A2", "A2"]]
    assert latest_sheet["cell_tags"] == [["header", "header"], ["data", "data"]]


def test_confirm_structure_version_marks_task_confirmed_and_returns_confirmed_version(tmp_path) -> None:
    session_factory, override_get_db = _override_db_session(
        f"sqlite:///{(tmp_path / 'test.db').as_posix()}"
    )
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    workbook_path = tmp_path / "source.xlsx"
    _build_sample_workbook(workbook_path)
    task_id = _create_uploaded_task(session_factory, workbook_path)

    parse_response = client.post(f"/api/tasks/{task_id}/parse")
    assert parse_response.status_code == 200

    review_response = client.get(f"/api/tasks/{task_id}/review")
    assert review_response.status_code == 200
    sheet = review_response.json()["sheets"][0]

    save_response = client.post(
        f"/api/tasks/{task_id}/structure-versions",
        json={
            "base_structure_version": 0,
            "sheets": [
                {
                    "sheet_id": sheet["sheet_id"],
                    "sheet_name": sheet["sheet_name"],
                    "sheet_index": sheet["sheet_index"],
                    "row_count": sheet["row_count"],
                    "col_count": sheet["col_count"],
                    "is_hidden": sheet["is_hidden"],
                    "merge_ranges": sheet["merge_ranges"],
                    "aligned_grid": [["Header", "Header"], ["100", "200"]],
                    "aligned_cell_roles": [["dimension", "dimension"], ["measure", "measure"]],
                    "aligned_source_map": [["A1", "A1"], ["A2", "B2"]],
                    "cell_tags": [["header", "header"], ["data", "data"]],
                }
            ],
        },
    )
    assert save_response.status_code == 201

    confirm_response = client.post(
        f"/api/tasks/{task_id}/confirm",
        json={"structure_version": 1},
    )
    app.dependency_overrides.clear()

    assert confirm_response.status_code == 200
    payload = confirm_response.json()
    assert payload["task_id"] == task_id
    assert payload["status"] == "confirmed"
    assert payload["structure_version"] == 1
    assert payload["confirmed_structure_version"] == 1

    with session_factory() as session:
        task = session.scalar(select(TaskRecordModel).where(TaskRecordModel.id == task_id))

    assert task is not None
    assert task.status == "confirmed"


def test_review_prefers_confirmed_structure_version_over_newer_unconfirmed_draft(tmp_path) -> None:
    session_factory, override_get_db = _override_db_session(
        f"sqlite:///{(tmp_path / 'test.db').as_posix()}"
    )
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    workbook_path = tmp_path / "source.xlsx"
    _build_sample_workbook(workbook_path)
    task_id = _create_uploaded_task(session_factory, workbook_path)

    assert client.post(f"/api/tasks/{task_id}/parse").status_code == 200
    review_response = client.get(f"/api/tasks/{task_id}/review")
    assert review_response.status_code == 200
    sheet = review_response.json()["sheets"][0]

    version_one = client.post(
        f"/api/tasks/{task_id}/structure-versions",
        json={
            "base_structure_version": 0,
            "sheets": [
                {
                    "sheet_id": sheet["sheet_id"],
                    "sheet_name": sheet["sheet_name"],
                    "sheet_index": sheet["sheet_index"],
                    "row_count": sheet["row_count"],
                    "col_count": sheet["col_count"],
                    "is_hidden": sheet["is_hidden"],
                    "merge_ranges": ["A1:B1"],
                    "aligned_grid": [["Header", "Header"], ["North", "South"]],
                    "aligned_cell_roles": [["dimension", "dimension"], ["dimension", "dimension"]],
                    "aligned_source_map": [["A1", "A1"], ["A2", "B2"]],
                    "cell_tags": [["header", "header"], ["data", "data"]],
                }
            ],
        },
    )
    assert version_one.status_code == 201
    assert client.post(f"/api/tasks/{task_id}/confirm", json={"structure_version": 1}).status_code == 200

    version_two = client.post(
        f"/api/tasks/{task_id}/structure-versions",
        json={
            "base_structure_version": 1,
            "sheets": [
                {
                    "sheet_id": sheet["sheet_id"],
                    "sheet_name": sheet["sheet_name"],
                    "sheet_index": sheet["sheet_index"],
                    "row_count": sheet["row_count"],
                    "col_count": sheet["col_count"],
                    "is_hidden": sheet["is_hidden"],
                    "merge_ranges": ["A1:B1", "A2:B2"],
                    "aligned_grid": [["Header", "Header"], ["Draft", "Draft"]],
                    "aligned_cell_roles": [["dimension", "dimension"], ["dimension", "dimension"]],
                    "aligned_source_map": [["A1", "A1"], ["A2", "A2"]],
                    "cell_tags": [["header", "header"], ["data", "data"]],
                }
            ],
        },
    )
    assert version_two.status_code == 201

    latest_review = client.get(f"/api/tasks/{task_id}/review")
    app.dependency_overrides.clear()

    assert latest_review.status_code == 200
    payload = latest_review.json()
    assert payload["status"] == "confirmed"
    assert payload["structure_version"] == 1
    assert payload["editable_structure_version"] == 2
    assert payload["sheets"][0]["aligned_grid"] == [["Header", "Header"], ["North", "South"]]


def test_review_recomputes_header_parsing_from_saved_structure_version(tmp_path) -> None:
    session_factory, override_get_db = _override_db_session(
        f"sqlite:///{(tmp_path / 'test.db').as_posix()}"
    )
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    workbook_path = tmp_path / "source.xlsx"
    _build_sample_workbook(workbook_path)
    task_id = _create_uploaded_task(session_factory, workbook_path)

    assert client.post(f"/api/tasks/{task_id}/parse").status_code == 200
    review_response = client.get(f"/api/tasks/{task_id}/review")
    assert review_response.status_code == 200
    sheet = review_response.json()["sheets"][0]

    save_response = client.post(
        f"/api/tasks/{task_id}/structure-versions",
        json={
            "base_structure_version": 0,
            "sheets": [
                {
                    "sheet_id": sheet["sheet_id"],
                    "sheet_name": sheet["sheet_name"],
                    "sheet_index": sheet["sheet_index"],
                    "row_count": sheet["row_count"],
                    "col_count": sheet["col_count"],
                    "is_hidden": sheet["is_hidden"],
                    "merge_ranges": ["A1:B1"],
                    "aligned_grid": [["Q1", "Q1"], ["Revenue", "Cost"]],
                    "aligned_cell_roles": [["dimension", "dimension"], ["measure", "measure"]],
                    "aligned_source_map": [["A1", "A1"], ["A2", "B2"]],
                    "cell_tags": [["header", "header"], ["data", "data"]],
                }
            ],
        },
    )
    assert save_response.status_code == 201

    latest_review = client.get(f"/api/tasks/{task_id}/review")
    app.dependency_overrides.clear()

    assert latest_review.status_code == 200
    payload = latest_review.json()
    sheet_payload = payload["sheets"][0]
    assert sheet_payload["aligned_grid"] == [["Q1", "Q1"], ["Revenue", "Cost"]]
    assert sheet_payload["header_row_span"] == 1
    assert sheet_payload["column_paths"] == [["Q1"], ["Q1"]]
    assert sheet_payload["column_kinds"] == ["dimension", "dimension"]


def test_review_does_not_treat_first_data_row_as_header_when_only_row_dimension_continues(tmp_path) -> None:
    session_factory, override_get_db = _override_db_session(
        f"sqlite:///{(tmp_path / 'test.db').as_posix()}"
    )
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    workbook_path = tmp_path / "row-dimension.xlsx"
    _build_row_dimension_workbook(workbook_path)
    task_id = _create_uploaded_task(session_factory, workbook_path)

    parse_response = client.post(f"/api/tasks/{task_id}/parse")
    assert parse_response.status_code == 200

    response = client.get(f"/api/tasks/{task_id}/review")
    app.dependency_overrides.clear()

    assert response.status_code == 200
    sheet = response.json()["sheets"][0]
    assert sheet["aligned_grid"] == [["Region", "Revenue"], ["East", "120"]]
    assert sheet["aligned_cell_roles"] == [["dimension", "dimension"], ["dimension", "measure"]]
    assert sheet["header_row_span"] == 1
    assert sheet["column_paths"] == [["Region"], ["Revenue"]]



def test_infer_formulas_requires_confirmed_task(tmp_path) -> None:
    session_factory, override_get_db = _override_db_session(
        f"sqlite:///{(tmp_path / 'test.db').as_posix()}"
    )
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    workbook_path = tmp_path / "source.xlsx"
    _build_sample_workbook(workbook_path)
    task_id = _create_uploaded_task(session_factory, workbook_path)

    assert client.post(f"/api/tasks/{task_id}/parse").status_code == 200

    response = client.post(f"/api/tasks/{task_id}/infer-formulas", json={})
    app.dependency_overrides.clear()

    assert response.status_code == 409
    assert response.json()["detail"] == "Task must be confirmed before formula inference"


def test_infer_formulas_returns_empty_when_llm_unavailable(tmp_path, monkeypatch) -> None:
    session_factory, override_get_db = _override_db_session(
        f"sqlite:///{(tmp_path / 'test.db').as_posix()}"
    )
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    monkeypatch.setattr(
        "app.services.formula.llm_formula_client.run_formula_inference",
        lambda **_: {"bad": "payload"},
    )

    workbook_path = tmp_path / "source.xlsx"
    _build_sample_workbook(workbook_path)
    task_id = _create_uploaded_task(session_factory, workbook_path)

    assert client.post(f"/api/tasks/{task_id}/parse").status_code == 200

    review = client.get(f"/api/tasks/{task_id}/review")
    assert review.status_code == 200
    sheet = review.json()["sheets"][0]

    assert client.post(
        f"/api/tasks/{task_id}/structure-versions",
        json={
            "base_structure_version": 0,
            "sheets": [
                {
                    "sheet_id": sheet["sheet_id"],
                    "sheet_name": sheet["sheet_name"],
                    "sheet_index": sheet["sheet_index"],
                    "row_count": sheet["row_count"],
                    "col_count": sheet["col_count"],
                    "is_hidden": sheet["is_hidden"],
                    "merge_ranges": sheet["merge_ranges"],
                    "aligned_grid": sheet["aligned_grid"],
                    "aligned_cell_roles": sheet["aligned_cell_roles"],
                    "aligned_source_map": sheet["aligned_source_map"],
                    "cell_tags": sheet.get("cell_tags", []),
                }
            ],
        },
    ).status_code == 201

    assert client.post(
        f"/api/tasks/{task_id}/confirm",
        json={"structure_version": 1},
    ).status_code == 200

    response = client.post(f"/api/tasks/{task_id}/infer-formulas", json={})
    app.dependency_overrides.clear()

    assert response.status_code == 200
    payload = response.json()
    assert payload["task_id"] == task_id
    assert payload["accepted_rules"] == []
    assert payload["rejected_count"] == 0


def test_formula_rules_endpoint_returns_filtered_rules(tmp_path) -> None:
    session_factory, override_get_db = _override_db_session(
        f"sqlite:///{(tmp_path / 'test.db').as_posix()}"
    )
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    workbook_path = tmp_path / "source.xlsx"
    _build_sample_workbook(workbook_path)
    task_id = _create_uploaded_task(session_factory, workbook_path)

    assert client.post(f"/api/tasks/{task_id}/parse").status_code == 200

    review = client.get(f"/api/tasks/{task_id}/review")
    assert review.status_code == 200
    sheet = review.json()["sheets"][0]

    assert client.post(
        f"/api/tasks/{task_id}/structure-versions",
        json={
            "base_structure_version": 0,
            "sheets": [
                {
                    "sheet_id": sheet["sheet_id"],
                    "sheet_name": sheet["sheet_name"],
                    "sheet_index": sheet["sheet_index"],
                    "row_count": sheet["row_count"],
                    "col_count": sheet["col_count"],
                    "is_hidden": sheet["is_hidden"],
                    "merge_ranges": sheet["merge_ranges"],
                    "aligned_grid": sheet["aligned_grid"],
                    "aligned_cell_roles": sheet["aligned_cell_roles"],
                    "aligned_source_map": sheet["aligned_source_map"],
                    "cell_tags": sheet.get("cell_tags", []),
                }
            ],
        },
    ).status_code == 201
    assert client.post(f"/api/tasks/{task_id}/confirm", json={"structure_version": 1}).status_code == 200

    from app.db.models.formula_rule_record import FormulaRuleRecordModel

    with session_factory() as session:
        session.add(
            FormulaRuleRecordModel(
                task_id=task_id,
                sheet_id=sheet["sheet_id"],
                formula_text="col_B = col_A + 10",
                formula_type="column_arithmetic",
                description="rule a",
                confidence=0.9,
                verification_passed=True,
                verification_score=0.8,
                prompt_version="day13_v1",
                model_name="mock/day13",
            )
        )
        session.add(
            FormulaRuleRecordModel(
                task_id=task_id,
                sheet_id=sheet["sheet_id"],
                formula_text="col_B = col_A + 10",
                formula_type="column_arithmetic",
                description="rule a dup",
                confidence=0.5,
                verification_passed=True,
                verification_score=0.7,
                prompt_version="day13_v1",
                model_name="mock/day13",
            )
        )
        session.commit()

    response = client.get(f"/api/tasks/{task_id}/formula-rules")
    app.dependency_overrides.clear()

    assert response.status_code == 200
    payload = response.json()
    assert payload["total_inferred"] == 2
    assert payload["passed"] == 1
    assert payload["filtered"] == 1
    assert payload["conflict"] == 0


def test_formula_rules_has_gap_when_no_passing_rules(tmp_path) -> None:
    session_factory, override_get_db = _override_db_session(
        f"sqlite:///{(tmp_path / 'test.db').as_posix()}"
    )
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    workbook_path = tmp_path / "source.xlsx"
    _build_sample_workbook(workbook_path)
    task_id = _create_uploaded_task(session_factory, workbook_path)

    assert client.post(f"/api/tasks/{task_id}/parse").status_code == 200
    review = client.get(f"/api/tasks/{task_id}/review")
    sheet = review.json()["sheets"][0]

    assert client.post(
        f"/api/tasks/{task_id}/structure-versions",
        json={
            "base_structure_version": 0,
            "sheets": [
                {
                    "sheet_id": sheet["sheet_id"],
                    "sheet_name": sheet["sheet_name"],
                    "sheet_index": sheet["sheet_index"],
                    "row_count": sheet["row_count"],
                    "col_count": sheet["col_count"],
                    "is_hidden": sheet["is_hidden"],
                    "merge_ranges": sheet["merge_ranges"],
                    "aligned_grid": sheet["aligned_grid"],
                    "aligned_cell_roles": sheet["aligned_cell_roles"],
                    "aligned_source_map": sheet["aligned_source_map"],
                    "cell_tags": sheet.get("cell_tags", []),
                }
            ],
        },
    ).status_code == 201
    assert client.post(f"/api/tasks/{task_id}/confirm", json={"structure_version": 1}).status_code == 200

    from app.db.models.formula_rule_record import FormulaRuleRecordModel

    with session_factory() as session:
        session.add(
            FormulaRuleRecordModel(
                task_id=task_id,
                sheet_id=sheet["sheet_id"],
                formula_text="col_B = col_A + 10",
                formula_type="column_arithmetic",
                description="low quality",
                confidence=0.1,
                verification_passed=True,
                verification_score=0.05,
                prompt_version="day13_v1",
                model_name="mock/day13",
            )
        )
        session.commit()

    response = client.get(f"/api/tasks/{task_id}/formula-rules")
    app.dependency_overrides.clear()

    assert response.status_code == 200
    payload = response.json()
    assert payload["has_gap"] is True
    assert payload["passed"] == 0


def test_acknowledge_gap_returns_success_for_gapped_task(tmp_path) -> None:
    session_factory, override_get_db = _override_db_session(
        f"sqlite:///{(tmp_path / 'test.db').as_posix()}"
    )
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    workbook_path = tmp_path / "source.xlsx"
    _build_sample_workbook(workbook_path)
    task_id = _create_uploaded_task(session_factory, workbook_path)

    assert client.post(f"/api/tasks/{task_id}/parse").status_code == 200
    review = client.get(f"/api/tasks/{task_id}/review")
    sheet = review.json()["sheets"][0]

    assert client.post(
        f"/api/tasks/{task_id}/structure-versions",
        json={
            "base_structure_version": 0,
            "sheets": [
                {
                    "sheet_id": sheet["sheet_id"],
                    "sheet_name": sheet["sheet_name"],
                    "sheet_index": sheet["sheet_index"],
                    "row_count": sheet["row_count"],
                    "col_count": sheet["col_count"],
                    "is_hidden": sheet["is_hidden"],
                    "merge_ranges": sheet["merge_ranges"],
                    "aligned_grid": sheet["aligned_grid"],
                    "aligned_cell_roles": sheet["aligned_cell_roles"],
                    "aligned_source_map": sheet["aligned_source_map"],
                    "cell_tags": sheet.get("cell_tags", []),
                }
            ],
        },
    ).status_code == 201
    assert client.post(f"/api/tasks/{task_id}/confirm", json={"structure_version": 1}).status_code == 200

    from app.db.models.formula_rule_record import FormulaRuleRecordModel

    with session_factory() as session:
        session.add(
            FormulaRuleRecordModel(
                task_id=task_id,
                sheet_id=sheet["sheet_id"],
                formula_text="col_B = col_A + 10",
                formula_type="column_arithmetic",
                description="low quality",
                confidence=0.1,
                verification_passed=True,
                verification_score=0.05,
                prompt_version="day13_v1",
                model_name="mock/day13",
            )
        )
        session.commit()

    response = client.post(
        f"/api/tasks/{task_id}/formula-rules/acknowledge-gap",
        json={"acknowledged": True},
    )
    app.dependency_overrides.clear()

    assert response.status_code == 200
    payload = response.json()
    assert payload["acknowledged"] is True
    assert "gap acknowledged" in payload["message"]

    # Verify task status was updated
    from sqlalchemy import select as sel

    from app.db.models.task_record import TaskRecordModel
    with session_factory() as session:
        task = session.scalar(
            sel(TaskRecordModel).where(TaskRecordModel.id == task_id),
        )
        assert task is not None
        assert task.status == "formula_gap_acknowledged"


