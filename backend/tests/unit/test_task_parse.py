from collections.abc import Generator
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import create_engine, event, func, select
from sqlalchemy.orm import Session, sessionmaker

from app.core.config import settings
from app.db.base import Base
from app.db.models.file_record import FileRecordModel
from app.db.models.task_record import TaskRecordModel
from app.db.session import get_db
from app.main import app


def _override_db_session(database_url: str) -> tuple[sessionmaker[Session], object]:
    engine = create_engine(
        database_url,
        future=True,
        connect_args={"check_same_thread": False},
    )

    @event.listens_for(engine, "connect")
    def _set_sqlite_pragma(dbapi_connection, connection_record) -> None:
        cursor = dbapi_connection.cursor()
        cursor.execute("PRAGMA foreign_keys=ON")
        cursor.close()

    Base.metadata.create_all(bind=engine)
    testing_session = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)

    def override_get_db() -> Generator[Session, None, None]:
        db = testing_session()
        try:
            yield db
        finally:
            db.close()

    return testing_session, override_get_db


def _build_sample_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "经营分析"
    sheet["A1"] = "Header"
    sheet.merge_cells("A1:B1")
    sheet["A2"] = 100
    sheet["B2"] = 200

    hidden_sheet = workbook.create_sheet("隐藏Sheet")
    hidden_sheet.sheet_state = "hidden"
    hidden_sheet["A1"] = "hidden"

    workbook.save(path)


def test_parse_task_reads_workbook_and_persists_sheets_and_cells(tmp_path) -> None:
    session_factory, override_get_db = _override_db_session(
        f"sqlite:///{(tmp_path / 'test.db').as_posix()}"
    )
    app.dependency_overrides[get_db] = override_get_db
    settings.upload_dir = str(tmp_path / "uploads")
    client = TestClient(app)

    workbook_path = tmp_path / "source.xlsx"
    _build_sample_workbook(workbook_path)

    with session_factory() as session:
        file_record = FileRecordModel(
            file_name="source.xlsx",
            file_path=str(workbook_path),
            file_type=".xlsx",
        )
        session.add(file_record)
        session.flush()
        task_record = TaskRecordModel(file_id=file_record.id, status="uploaded")
        session.add(task_record)
        session.commit()
        session.refresh(task_record)
        task_id = task_record.id

    response = client.post(f"/api/tasks/{task_id}/parse")
    app.dependency_overrides.clear()

    assert response.status_code == 200
    payload = response.json()
    assert payload["task_id"] == task_id
    assert payload["status"] == "waiting_confirm"
    assert len(payload["sheets"]) == 2
    assert payload["sheets"][0]["sheet_name"] == "经营分析"
    assert payload["sheets"][0]["row_count"] == 2
    assert payload["sheets"][0]["col_count"] == 2

    with session_factory() as session:
        task = session.scalar(select(TaskRecordModel).where(TaskRecordModel.id == task_id))
        assert task is not None
        assert task.status == "waiting_confirm"

        sheet_rows = session.execute(
            select(func.count()).select_from(Base.metadata.tables["sheets"])
        ).scalar_one()
        cell_rows = session.execute(
            select(func.count()).select_from(Base.metadata.tables["cells"])
        ).scalar_one()

        assert sheet_rows == 2
        assert cell_rows == 5

        merged_cells = session.execute(
            select(
                Base.metadata.tables["cells"].c.address,
                Base.metadata.tables["cells"].c.merge_range,
            ).where(Base.metadata.tables["cells"].c.merge_range.is_not(None))
        ).all()

    assert ("A1", "A1:B1") in merged_cells
    assert ("B1", "A1:B1") in merged_cells


def test_parse_returns_404_for_missing_task(tmp_path) -> None:
    _, override_get_db = _override_db_session(f"sqlite:///{(tmp_path / 'test.db').as_posix()}")
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    response = client.post("/api/tasks/999/parse")
    app.dependency_overrides.clear()

    assert response.status_code == 404
    assert response.json()["detail"] == "Task not found"


def test_parse_returns_404_when_uploaded_file_is_missing(tmp_path) -> None:
    session_factory, override_get_db = _override_db_session(
        f"sqlite:///{(tmp_path / 'test.db').as_posix()}"
    )
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    missing_workbook_path = tmp_path / "missing.xlsx"
    with session_factory() as session:
        file_record = FileRecordModel(
            file_name="missing.xlsx",
            file_path=str(missing_workbook_path),
            file_type=".xlsx",
        )
        session.add(file_record)
        session.flush()
        task_record = TaskRecordModel(file_id=file_record.id, status="uploaded")
        session.add(task_record)
        session.commit()
        session.refresh(task_record)
        task_id = task_record.id

    response = client.post(f"/api/tasks/{task_id}/parse")
    app.dependency_overrides.clear()

    assert response.status_code == 404
    assert response.json()["detail"] == "Uploaded file not found"


def test_parse_replaces_previous_sheet_and_cell_rows_for_same_task(tmp_path) -> None:
    session_factory, override_get_db = _override_db_session(
        f"sqlite:///{(tmp_path / 'test.db').as_posix()}"
    )
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    workbook_path = tmp_path / "source.xlsx"
    _build_sample_workbook(workbook_path)

    with session_factory() as session:
        file_record = FileRecordModel(
            file_name="source.xlsx",
            file_path=str(workbook_path),
            file_type=".xlsx",
        )
        session.add(file_record)
        session.flush()
        task_record = TaskRecordModel(file_id=file_record.id, status="uploaded")
        session.add(task_record)
        session.commit()
        session.refresh(task_record)
        task_id = task_record.id

    first_response = client.post(f"/api/tasks/{task_id}/parse")
    assert first_response.status_code == 200

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "经营分析"
    sheet["A1"] = "Updated"
    sheet["A2"] = 300
    workbook.save(workbook_path)
    workbook.close()

    second_response = client.post(f"/api/tasks/{task_id}/parse")
    app.dependency_overrides.clear()

    assert second_response.status_code == 200

    with session_factory() as session:
        sheet_rows = session.execute(
            select(func.count()).select_from(Base.metadata.tables["sheets"])
        ).scalar_one()
        cell_rows = session.execute(
            select(func.count()).select_from(Base.metadata.tables["cells"])
        ).scalar_one()
        raw_values = session.execute(
            select(Base.metadata.tables["cells"].c.raw_value).order_by(Base.metadata.tables["cells"].c.id)
        ).scalars().all()

    assert sheet_rows == 1
    assert cell_rows == 2
    assert raw_values == ["Updated", "300"]


def test_parse_persists_hidden_sheet_metadata(tmp_path) -> None:
    session_factory, override_get_db = _override_db_session(
        f"sqlite:///{(tmp_path / 'test.db').as_posix()}"
    )
    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    workbook_path = tmp_path / "source.xlsx"
    _build_sample_workbook(workbook_path)

    with session_factory() as session:
        file_record = FileRecordModel(
            file_name="source.xlsx",
            file_path=str(workbook_path),
            file_type=".xlsx",
        )
        session.add(file_record)
        session.flush()
        task_record = TaskRecordModel(file_id=file_record.id, status="uploaded")
        session.add(task_record)
        session.commit()
        session.refresh(task_record)
        task_id = task_record.id

    response = client.post(f"/api/tasks/{task_id}/parse")
    app.dependency_overrides.clear()

    assert response.status_code == 200

    with session_factory() as session:
        sheet_rows = session.execute(
            select(
                Base.metadata.tables["sheets"].c.sheet_name,
                Base.metadata.tables["sheets"].c.sheet_index,
                Base.metadata.tables["sheets"].c.is_hidden,
            ).order_by(Base.metadata.tables["sheets"].c.sheet_index)
        ).all()

    assert sheet_rows == [("经营分析", 0, False), ("隐藏Sheet", 1, True)]
