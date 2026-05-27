from pathlib import Path

from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.db.models.cell_record import CellRecordModel
from app.db.models.sheet_record import SheetRecordModel
from app.db.models.task_record import TaskRecordModel


def _build_merge_lookup(worksheet) -> dict[str, str]:
    merge_lookup: dict[str, str] = {}
    for merged_range in worksheet.merged_cells.ranges:
        range_string = str(merged_range)
        for row in worksheet[range_string]:
            for cell in row:
                merge_lookup[cell.coordinate] = range_string
    return merge_lookup


def parse_task_workbook(db: Session, task_id: int) -> dict[str, object]:
    task = db.scalar(select(TaskRecordModel).where(TaskRecordModel.id == task_id))
    if task is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Task not found")

    file_record = task.file
    workbook_path = Path(file_record.file_path)
    if not workbook_path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Uploaded file not found")

    workbook = load_workbook(workbook_path, data_only=False)

    try:
        existing_sheet_ids = db.scalars(
            select(SheetRecordModel.id).where(SheetRecordModel.task_id == task.id)
        ).all()
        if existing_sheet_ids:
            db.execute(delete(CellRecordModel).where(CellRecordModel.sheet_id.in_(existing_sheet_ids)))
            db.execute(delete(SheetRecordModel).where(SheetRecordModel.task_id == task.id))
            db.flush()

        parsed_sheets: list[dict[str, int | str]] = []
        for index, worksheet in enumerate(workbook.worksheets):
            sheet_record = SheetRecordModel(
                task_id=task.id,
                sheet_name=worksheet.title,
                sheet_index=index,
                row_count=worksheet.max_row,
                col_count=worksheet.max_column,
                is_hidden=worksheet.sheet_state != "visible",
            )
            db.add(sheet_record)
            db.flush()

            merge_lookup = _build_merge_lookup(worksheet)
            cell_records: list[CellRecordModel] = []
            for row in worksheet.iter_rows():
                for cell in row:
                    raw_value = None if cell.value is None else str(cell.value)
                    merge_range = merge_lookup.get(cell.coordinate)
                    cell_records.append(
                        CellRecordModel(
                            sheet_id=sheet_record.id,
                            row_index=cell.row,
                            col_index=cell.column,
                            address=cell.coordinate,
                            raw_value=raw_value,
                            normalized_value=raw_value,
                            value_type=getattr(cell, "data_type", "unknown"),
                            is_merged=merge_range is not None,
                            merge_range=merge_range,
                        )
                    )

            db.add_all(cell_records)
            parsed_sheets.append(
                {
                    "sheet_id": sheet_record.id,
                    "sheet_name": sheet_record.sheet_name,
                    "row_count": sheet_record.row_count,
                    "col_count": sheet_record.col_count,
                }
            )

        task.status = "waiting_confirm"
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        workbook.close()

    return {
        "task_id": task.id,
        "status": task.status,
        "sheets": parsed_sheets,
    }
